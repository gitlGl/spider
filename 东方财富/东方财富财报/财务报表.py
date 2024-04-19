import requests,os
import time, signal,sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#os.chdir(sys.path[0])
current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  

表字段 = {"利润表":(('REPORT_DATE', '报告期'), ('PARENT_NETPROFIT', '净利润(元)'),
 ('PARENT_NETPROFIT_RATIO', '净利润同比(%)'), ('DEDUCT_PARENT_NETPROFIT', '扣非归母净利润(元)'),
  ('DPN_RATIO', '扣非归母净利润同比(%)'), ('INTEREST_NI', '利息净收入(元)'), 
  ('FEE_COMMISSION_NI', '手续费及佣金净收入(元)'), ('FCN_RATIO', '手续费及佣金净收入同比(%)'), 
  ('TOTAL_OPERATE_INCOME', '营业总收入(元)'), ('TOI_RATIO','营业总收入同比(%)'),
  ('OPERATE_TAX_ADD', '营业税金及附加(元)'), ('MANAGE_EXPENSE', '管理费用(元)'), 
  ('TOTAL_OPERATE_COST', '营业总支出(元)'), ('TOE_RATIO', '营业总支出同比(%)'), 
  ('OPERATE_PROFIT', '营业利润(元)'), ('OPERATE_PROFIT_RATIO', '营业利润同比(%)'), 
  ('TOTAL_PROFIT', '利润总额(元)'), ('NOTICE_DATE', '公告日期')),


  "资产负债表":(('REPORT_DATE',"报告期"),("TOTAL_ASSETS","总资产(元)"),("TOTAL_ASSETS_RATIO","总资产同比(%"),
  ("MONETARYFUNDS","货币资金(元)"),("MONETARYFUNDS_RATIO","同比(%)"),
  ("SETTLE_EXCESS_RESERVE","结算备付金(元"),("SER_RATIO","同比(%)"),
  ("AVAILABLE_SALE_FINASSET","可供出售金融资产(元)"),("ASF_RATIO","同比(%)"),
  ("TOTAL_LIABILITIES","总负债(元)"),("TOTAL_LIAB_RATIO","总负债同比(%"),
  ("BORROW_FUND","拆入资金(元)"),("BORROW_FUND_RATIO","同比(%)"),("SELL_REPO_FINASSET","卖出回购金融资产款(元)"),
  ("SRF_RATIO","同比(%)"),("AGENT_TRADE_SECURITY","代理买卖证券款(元"),
  ("ATS_RATIO","同比(%)"),("TOTAL_EQUITY","股东权益合计(元"),
  ("TOTAL_EQUITY_RATIO","股东权益同比(%)"),("DEBT_ASSET_RATIO","资产负债率(%"),('NOTICE_DATE',"公告日期")),


  "现金流量表":(('REPORT_DATE', '报告期'),("CCE_ADD","净现金流(元)"),("CCE_ADD_RATIO","净现金流同比(%)"),
  ("NETCASH_OPERATE","经营性现金流量净额(元"),("NETCASH_OPERATE_RATIO","经营性现金流量净额占比(%)"),
  ("RECEIVE_INTEREST_COMMISSION","金额(元)"),("RIC_RATIO","占比(%)"),("NETCASH_INVEST","投资性现金流量净额(元)"),
  ("NETCASH_INVEST_RATIO","投资性现金流量净额净额占比(%)"),("RECEIVE_INVEST_INCOME","金额(元)"),("RII_RATIO","占比(%)"),
  ("INVEST_PAY_CASH","金额(元)"),("IPC_RATIO","占比(%)"),("NETCASH_FINANCE","融资性现金流量净额(元)"),
  ("NETCASH_FINANCE_RATIO","融资性现金流量净额占比(%)"),('NOTICE_DATE', '公告日期')),

  "业绩报表":(
    ("REPORTDATE","报告期"),("BASIC_EPS","每股收益(元"),("DEDUCT_BASIC_EPS","每股收益(扣除)(元)"),
    ("TOTAL_OPERATE_INCOME","营业总收入(元)"),("YSTZ","同比增长(%)"),
    ("YSHZ","季度环比增长(%)"),("PARENT_NETPROFIT","净利润(元)"),( "SJLTZ","同比增长(%)"),("SJLHZ","季度环比增长(%)"),
    ("BPS","每股净资产(元)"),("WEIGHTAVG_ROE","净资产收益率(%)"),("MGJYXJJE","每股经营现金流量(元)"),("XSMLL","销售毛利率(%)"),("ASSIGNDSCRPT","利润分配"),
    ("ZXGXL","股息率(%)"),("NOTICE_DATE","首次公告日期"),("UPDATE_DATE","最新公告日期")

  )
  }

公司名 = ("SECURITY_NAME_ABBR","SECURITY_CODE")

# [item[1] for item in 表字段["业绩报表"] ]
# book = load_workbook("tes.xlsx")
# sheet = book["业绩报表"]
# for column_num, value in enumerate([item[1] for item in 表字段["业绩报表"] ], start=1):
#     sheet.cell(row=1, column=column_num, value=value)
# book.save("tes.xlsx")
# book.close()

def check(number):#检查xls文件格式，调整文件内容
    if type(number) == str:
        if len(number) == 6 and number.isdigit():
            return number
    if type(number) == float:
        tem = str(int(number))
        if len(tem) < 6:
            lenth = 6 - len(tem)
            for i in range(lenth):
                tem = "0" + tem
        return tem
    else:
        print("格式错误：",number)
def getNumber():#获取xls文件内的公司代码
    # 加载 Excel 文件
    workbook = load_workbook(file_name_xls)
    # 选择第一个工作表
    sheet = workbook.active
    return [cell.value for cell in sheet['A'] if check(cell.value) ]


def get_data_json(number):
    number = '"' + number + '"'

    if 财报表类型元组[1]  == "业绩报表":
        filter = '(SECURITY_CODE%3D' + number +')' +  '(DATEMMDD%3D' + 业绩报表财报类型 + ')'
        url = f"""https://datacenter-web.eastmoney.com/api/data/v1/get?
            &sortTypes=-1&columns=ALL&filter={filter}&reportName={财报表类型元组[0]}"""

    else:
        filter = '(SECURITY_CODE%3D' + number +')' +  '(DATE_TYPE_CODE%3D' + 财报类型 + ')'
        url = f"""https://datacenter-web.eastmoney.com/api/data/v1/get?
    &sortColumns=REPORT_DATE&sortTypes=-1&columns=ALL&filter={filter}&reportName={财报表类型元组[0]}"""
    headers =  {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"}
        # 发起请求

    print(url)
    while True:
        try:
            with session.get(url) as req: 
                data_json =  req.json()
                result = data_json['result']
                if result is None:
                    return None
                break   
        except Exception as e:
           print("请求失败，稍后重试")
           time.sleep(60)
    return result["data"]

def convert(field_data,data_temp):
    type_ = type(field_data)
    if type_ is float or type_  is int:
        if 4 < len(str(int(field_data))) < 9:
            num = round(field_data/10000,2)

            data_temp.append(str(num)+"万")
        elif len(str(int(field_data))) > 8:

            num = round(field_data/100000000,2)
            data_temp.append(str(num)+"亿")
        else:
            num = round(field_data,2)

            data_temp.append(str(num))
    if type_ is str:
        if "00:00:00" in field_data:
            field_data = field_data[:10]
        data_temp.append(field_data)

def get_data(number):
    if number + 财报表类型元组[1] in download_Progress:
            print(number + 财报表类型元组[1]+"已下载")
            return
    data = []
    data_json = get_data_json(number)
    if not data_json:
        print(f"企业{number}没有数据")
        return
    for item in data_json:
        data_temp = []
        if 财报表类型元组[1]  == "业绩报表":
            data_temp.append(业绩报表财报类型.strip('"'))
        else: data_temp.append(财报类型字典[item["DATE_TYPE_CODE"]])
        data_temp.append(item["SECURITY_NAME_ABBR"] +"（" +item["SECURITY_CODE"] + "）")

        for field,field_chines in 表字段[财报表类型元组[1]]:
            if  item[field] is  None:
                data_temp.append(None)

            elif r"%" in field_chines :

                if abs(item[field]) < 10:
                    num = round(item[field],3)
                    data_temp.append(str(num))
                else:
                    num = round(item[field],2)
                    data_temp.append(str(num))     
            else:
                    convert(item[field],data_temp)

        data.append(data_temp)
    return data


def main():
    list_number = getNumber()
    book = load_workbook(财务表路径)
    flag = 0#程序退出标志
    def signal_handler(sig, frame):
        while True:
            f = input("退出程序？输入Y/N:")
            if f == "Y" or f == 'y':
                nonlocal flag
                flag = 1
                break
            if f == "N" or f == 'n':
                break
        
        #print(sig,type(sig))#(type int)
        # 信号处理函数
        # stack = traceback.extract_stack(frame)
        # for file_name, lineno, function, _ in stack:
        #     print(file_name,lineno,function)
        # locals_dict = frame.f_locals
        # print("Local Variables:")
        # print(locals_dict["number"])
        # # for var_name, var_value in locals_dict.items():
        # #     print(f"{var_name} = {var_value}")
        # book.save(财务表路径)
        # book.close()
        # sys.exit(0)

    # 注册信号处理函数
    signal.signal(signal.SIGINT, signal_handler)

# 其他程序代码...
    for number in list_number:
        if flag:
            book.save(财务表路径)
            book.close()
            sys.exit(0)

        data = get_data(number)
        if not data:
            continue

        sheet = book[财报表类型元组[1]]#选择excel表

        # 将DataFrame的数据逐行写入工作表
        for row in data:
            sheet.append(row)

        # 保存Excel文件
        book.save(财务表路径)
        #记录下载进度
        with open(file_name, "a",encoding="utf8") as f:
            f.write(number + 财报表类型元组[1] + "\n")

    book.close()


def readTxt(file_name):# 读取爬取进度
    if not os.path.exists(file_name):
        with open(file_name, "w",encoding="utf8") as f:
            f.write('')
    with open(file_name, "r",encoding="utf8") as f:
        data = f.read().splitlines()
        return data   

财报类型字典 = {"001":"年报","002":"中报","003":"一季报","004":"三季报"}

业绩报表类型列表 = ["年报","一季报","半年报","三季报"]

file_name = "已下载公司代码.txt"#记录进度
download_Progress = readTxt(file_name)# 读取已爬取进度

"""
(("RPT_DMSK_FN_INCOME","利润表"),("RPT_DMSK_FN_CASHFLOW","现金流量表"),
("RPT_DMSK_FN_BALANCE","资产负债表"),("RPT_LICO_FN_CPD","业绩报表"))
"""
"""
目前支持下载利润表，现金流量表，资产负债表，业绩报表
需要下载那个，就在元组里增加对应元组。
也可以选择你要下载的财报类型，比如年报半年报，如果填写为"",则后默认下载所有类型（"年报","一季报","半年报","三季报"）

 
"""


财报表类型 = (("RPT_DMSK_FN_INCOME","利润表"),("RPT_DMSK_FN_CASHFLOW","现金流量表"),
("RPT_DMSK_FN_BALANCE","资产负债表"),("RPT_LICO_FN_CPD","业绩报表"))

session = requests.session()

财报表类型元组 :tuple = None
财报类型 = '"001"'
业绩报表财报类型 = '"年报"'
财务表路径 = "财务表.xlsx"
file_name_xls = "股票代码.xlsx"#需要下载的公司代码所在的xls文件,出口上市公司.xls

print("退出请按ctrl + C，请勿强制退出，否则导致数据损坏")

def countdown(seconds):
    for i in range(seconds, 0, -1):
        print(i, end='\r')
        time.sleep(1)
countdown(5)
for item in 财报表类型:
    财报表类型元组 = item
    main()
    session.close()