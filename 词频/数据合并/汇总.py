from openpyxl import load_workbook
import os
current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))

def check(number):#检查xls文件格式，调整文件内容
    if type(number) == str:
        tem = number
        if len(tem) < 6:
            lenth = 6 - len(tem)
            for i in range(lenth):
                tem = "0" + tem
        return tem
        
    if type(number) == float:
        tem = str(int(number))
        if len(tem) < 6:
            lenth = 6 - len(tem)
            for i in range(lenth):
                tem = "0" + tem
        return tem
    
    if type(number) == int:
        tem = str(int(number))
        if len(tem) < 6:
            lenth = 6 - len(tem)
            for i in range(lenth):
                tem = "0" + tem
        return tem
        
    else:
        print("格式错误：",number)
        
def get_po_num(local_file):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    sheet = workbook. active
    std_years = []
   
    for cell1,cell2 in zip(sheet["A"][1:],sheet["B"][1:]):
        value1 =  cell1.value
        value2 =  cell2.value 
        if value1 and value2 :
            std_years.append(
                (check(str(value1)),
                str(value2)[:4])
                )
            
        else:
           std_years.append(None,None)
        
   
    return std_years

def create_index_mapping(lst):
    index_mapping = {}
    for i, tup in enumerate(lst,start=2):
        index_mapping[tup] = i
    return index_mapping


def extract_row(ws, row_index):
    
    # 提取指定行的数据，从第三列开始
    data = []
    for cell in ws[row_index][2:]:  # 从第三列开始迭代
        data.append(cell.value)

    return data

def get_max_column(target):
    target_wb = load_workbook(filename=target)
    target_ws = target_wb.active
    col_idx = target_ws.max_column 
    target_wb.save(target)
    target_wb.close()
    return col_idx


def incrementally_write_data(target_excel,max_column, data):
    # 加载目标 Excel 文件
    target_wb = load_workbook(filename=target_excel)
    target_ws = target_wb.active
    
    # 写入数据到指定行的新列中
     # 新列的索引为当前最大列索引加一
    for  data_to_write, target_row in data:
        for column, cell_value in enumerate(data_to_write, start=1):
            target_ws.cell(row=target_row, column= column +max_column, value=cell_value)
      
    
    # 保存目标 Excel 文件
    target_wb.save(target_excel)
    target_wb.close()


def main(source,target):
   
    词频代码与年份列表 = get_po_num(target)
    index_mapping  = create_index_mapping(词频代码与年份列表)
    变量代码与年份列表 = get_po_num(source)
    max_column = get_max_column(target)
    wb = load_workbook(source)
    ws = wb.active
    column_name = extract_row(ws, 1) 
    incrementally_write_data(target,max_column, [(column_name,1)])
    

    count = 0
    data = []
    for 变量索引,代码与年份 in enumerate( 变量代码与年份列表,start=2):
    
        词频索引 = index_mapping.get(代码与年份, None)
        
        if not 词频索引:
            continue
        
        column_data = extract_row(ws, 变量索引)
        data.append((column_data,词频索引 ))
    
        if len(data) == 200:
            count = count + 1
            incrementally_write_data(target,max_column,  data)
            data = []
            print("进度：",count * 200)
            

    incrementally_write_data(target, max_column, data)
        
            
    wb.save(source)
    wb.close() 
        
        
if "__main__" == __name__:
   
    target = '词频汇总.xlsx'
    dirs = [x for x in os.listdir(".") if x.endswith(".xlsx")]
    dirs.remove('词频汇总.xlsx')
    for source in dirs:
        main(source,target)