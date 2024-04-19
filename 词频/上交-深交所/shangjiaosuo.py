
import requests,re,csv
import json ,time,copy

from openpyxl import load_workbook
import os
current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path)) 

def get_code_name_pairs(file_name):
    # 加载 Excel 文件
    wb = load_workbook(file_name)
    # 获取活动工作表
    ws = wb.active

    # 创建一个空字典，用于存储键值对
    code_name_pairs = {}

    # 遍历工作表的行，读取键值对数据
    for row in ws.iter_rows(values_only=True):
        code = row[0]  # 第一列为键
        name = row[1]  # 第二列为值
        code_name_pairs[code] = name 
    return code_name_pairs

code_name_pairs = get_code_name_pairs("code_name_pairs.xlsx")


def is_record(results):
    tem_results = copy.deepcopy(results)
    years = []
    for index1,result in enumerate(tem_results):
        title = result['TITLE']
        year = result['BULLETIN_YEAR']
        if result['BULLETIN_YEAR'] in years:
                continue
        if "摘要"  in title:
            results.remove(result)
            continue
            
        if "取消"  in title:
              results.remove(result)
              continue
        if "英文"  in title:
             results.remove(result)
             continue
        if '说明' in title:
             results.remove(result)
             continue
       
        if "修订" in  title or "更新" in  title or "更正" in  title:
             for index2, x in enumerate( tem_results):
                 if x['BULLETIN_YEAR'] == result['BULLETIN_YEAR'] and index1!=index2:
                     print("更新")
                     years.append(x['BULLETIN_YEAR'])
                     results.remove(x)
                
    return results
            
     


#str(int(year)+1)}-01-01~{str(int(year)+1)}-12-31
def disclosure(code):
        """
        获得该公司的股票代码、报告类型、年份、定期报告披露日期、定期报告pdf下载链接，返回DataFrame。
        :param code:  股票代码
        :return: 返回DataFrame
        """
     
        datas = []
       
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36',
            'Referer': 'http://www.sse.com.cn/disclosure/listedinfo/regular/'}
        
        base = 'http://query.sse.com.cn/security/stock/queryCompanyBulletin.do?isPagination=true&productId={code}&securityType=0101%2C120100%2C020100%2C020200%2C120200&reportType2=DQBG&reportType=YEARLY&beginDate={beginDate}&endDate={endDate}&pageHelp.pageSize=25&pageHelp.pageCount=50&pageHelp.pageNo=1&pageHelp.beginPage=1&pageHelp.cacheSize=1&pageHelp.endPage=5'
       
        
        for year in list_years:
            url = base.format(code=code, beginDate=f"{str(int(year)+1)}-01-01", endDate=f"{str(int(year)+1)}-12-31")
           
            print('正在获取{}    {}定期报告披露信息'.format(year,code))
            
            while True:
                try:
                    resp = requests.get(url, headers=headers, cookies=cookies)
                    break
                except:
                    print('获取失败，重新获取')
                    time.sleep(60)
          
            raw_data = json.loads(resp.text)
           
            results = raw_data['result']
           
            results = [x for x in results if "摘要" not in x['BULLETIN_TYPE'] ]
            results = is_record(results)
          
            for result in results:
                title = result['TITLE']
              
                pdf = 'http://www.sse.com.cn' + result['URL']
            
                company = code_name_pairs[str(code)]
                _type = result['BULLETIN_TYPE']
                year = result['BULLETIN_YEAR']
                date = result['SSEDATE']
                title = result['TITLE']
                data = [company, code, _type, year, date, pdf,title]
                datas.append(data)
                    
        return datas
        
def readTxt(file_name):# 读取已下载的公司代码
    if not os.path.exists(file_name):
        with open(file_name, "w") as f:
            f.write('')
    with open(file_name, "r") as f:
        data = f.read().splitlines()
        return data    

list_code = readTxt("上交所进度.txt")
list_years = ["2015","2016","2017","2018","2019","2020","2021","2022","2023"] # 下载所需要的年份年报
cookies = {"Cookie": 'ba17301551dcbaf9_gdp_session_id=fe0089fe-71b7-4375-80c5-2b80d625e1df; gdp_user_id=gioenc-7b87geg9%2C5463%2C53ec%2Ca8g7%2C41aadbc577b5; ba17301551dcbaf9_gdp_session_id_sent=fe0089fe-71b7-4375-80c5-2b80d625e1df; VISITED_MENU=%5B%229075%22%2C%2210766%22%5D; sseMenuSpecial=14887; ba17301551dcbaf9_gdp_sequence_ids={%22globalKey%22:94%2C%22VISIT%22:2%2C%22PAGE%22:22%2C%22VIEW_CLICK%22:68%2C%22CUSTOM%22:4%2C%22VIEW_CHANGE%22:2}'}

headers=['company','code', 'type', 'year', 'date', 'pdf',"title"]
for code in code_name_pairs.keys():
    if code  in list_code:
        print("{}已下载".format(code))
        continue
    else:
        print('正在获取{}'.format(code))
    
    datas = disclosure(code)
    time.sleep(0.5)
    
    if not os.path.exists('上交所.csv'):
        with open('上交所.csv','a+',encoding="utf8",newline='') as f:
            f_csv = csv.writer(f)
            f_csv.writerow(headers)
        
    with open('上交所.csv','a+',encoding="utf8",newline='') as f:
        f_csv = csv.writer(f)
        f_csv.writerows(datas)
        
    with open("上交所进度.txt", 'a+') as f:
        f.write(code + '\n')
    
    
