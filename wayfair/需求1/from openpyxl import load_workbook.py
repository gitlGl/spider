import os
from openpyxl import load_workbook
current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  
def read_excel_file(local_file,sheet_name):
    workbook = load_workbook(local_file)
    sheet = workbook[sheet_name]
    
    for row in sheet.iter_rows(values_only=True):
        yield row[0]

        
gen = read_excel_file("Testing11.20-11.30.xlsx","Testing11.20-11.30")

for x in gen:
    print(x)