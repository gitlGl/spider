import requests,json,csv,time,os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from decimal import Decimal
import random

current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  




local_file = "12-1000 测试.xlsx"
sheet_name = "12-1000"

voucher_number_file = "test.xlsx"

sheet_name_vo_num = "Sheet1"

payload = {
  "poNumber": "CS488670090",
  "isSap": "0"
}
#重要
cookies = {"Cookie":'CSNUtId=225641ce-65ba-becf-46b3-a7fa51c48d02; ExCSNUtId=225641ce-65ba-becf-46b3-a7fa51c48d02; partner_home_language=eng_US; _ga=GA1.3.938233253.1706737365; _pxvid=a7fce01d-c081-11ee-bb46-dbc7fc54ee7d; _gid=GA1.2.1417150947.1712381010; _gid=GA1.3.1417150947.1712381010; canary=0; _pxhd=TmsS60LOo31o0LDAxGHhsbNcgST4d37ltqRdB9t-cKpAEAxPTXo4YXR7-18Yj60eVlfFHs9TXAAA9lIsdS5XVQ==:TgM8zpeEKz1reCpgFEecovg/uRjQUW9lkmNu1pfsp8FawFVjfC4dgT42UMJPkn14DTusZPV9DwyL79fdMc0fIsTyc/m9YdpispscLY03q5M=; _px3=32901c5d7d1c0d7a750f14a26cac6ab6248d227d53d6b831cb7dc81b604a1eb1:HqSpvQcTBpxGbsk8cDAx2u11tRcaRUDuI7Xc6X5jT7yq+FvR52X9Qn8P/3vGC4+ox0sw0dWdWbUS9FUcOHjwKA==:1000:yZxadNHY0V6oY5C7+EJuzq6fdJAPQyA/uqAgz1SjceklQoYfgJvELzKlOo1x4aHN0wVy4vJlXenBnZqXd+0hS3Ryl1CThAx9Zfv+SUKor6GEF0JYbNcVHiYHkat9NgEOCe77FLIFbij3ZLUy1mvcPWq2gxUfgD8tWkVZJ26qMI2uPeI0YhnIrPrU/tHvglxcj2lm01KmUd+mll7zvyj/8ZvwFeUB+RVao93OLZyAmdk=; extranet_WFSID=bee80839b628ecaeff2ea065a6175cb4; _gat_UA-2081664-5=1; _gat_UA-260345920-1=1; CSNEXTUID=FEAEA1C7-52A8-4DAD-8BE0-0FD47C97D203; MEDIAHUBPREVIEW=1; supplierID=58079; featureDetect={"isTouch":false,"hasMQ":true,"deviceWidth":1447,"deviceHeight":1012,"devicePixelRatio":1.9250000715255737}; _ga_TLNRVZPPMM=GS1.2.1712468626.113.1.1712468642.0.0.0; _ga=GA1.2.938233253.1706737365; pxcts=d68b57b2-f4a1-11ee-8709-771d94cdbb3c; session_timeout=1712475848; bearer_token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc19hZG1pbiI6ZmFsc2UsImlzX2FkbWluX3ZpZXciOmZhbHNlLCJvcmdhbml6YXRpb25fYWNjb3VudF9uYW1lIjoiIiwiaXNfc3VwcGxpZXJfc2V0dXBfYWN0aXZlIjpmYWxzZSwibGFuZ3VhZ2VfaWQiOiJ6aF9jbiIsImxpc3Rfc2VwYXJhdG9yIjoiLCIsImRhdGVfZm9ybWF0X2lkIjoyLCJudW1iZXJfZm9ybWF0X2lkIjoxLCJleHRyYW5ldF91c2VyX2lkIjoxMTgxNDcsImV4dHJhbmV0X3VzZXJuYW1lIjoiY2luZHk4IiwiZXh0cmFuZXRfZmlyc3RfbmFtZSI6IkNpbmR5IiwiZXh0cmFuZXRfbGFzdF9uYW1lIjoiQ2luZHkiLCJleHRyYW5ldF9lbWFpbCI6ImxvdmV5eWNkQGdtYWlsLmNvbSIsImV4dHJhbmV0X3VzZXJfdHlwZV9pZCI6MSwiZXh0cmFuZXRfdXNlcl9zdGF0dXMiOjEsInNjb3BlIjoie1wiYXBwbGljYXRpb25cIjpbXCJyZWFkOnNpdGVfcmVhZHlfYXNzZXRzXCIsXCJ3cml0ZTpzaXRlX3JlYWR5X2Fzc2V0c1wiLFwicmVhZDozZF9tb2RlbHNcIixcIndyaXRlOjNkX21vZGVsc1wiLFwicmVhZDpub25fc2l0ZV9yZWFkeV9hc3NldHNcIixcIndyaXRlOm5vbl9zaXRlX3JlYWR5X2Fzc2V0c1wiLFwicmVhZDpwZXJpZ29sZF9yZXN0cmljdGVkXCIsXCJ3cml0ZTpwZXJpZ29sZF9yZXN0cmljdGVkXCIsXCJyZWFkOmhndHZfcmVzdHJpY3RlZFwiLFwid3JpdGU6aGd0dl9yZXN0cmljdGVkXCIsXCJyZWFkOnZpc3VhbF9zZWFyY2hfcmVzdHJpY3RlZFwiLFwid3JpdGU6dmlzdWFsX3NlYXJjaF9yZXN0cmljdGVkXCIsXCJhZG1pbjpkYW1cIixcInJlYWQ6bWVyY2hfY29tcGxpYW5jZV9yZXN0cmljdGVkXCIsXCJ3cml0ZTptZXJjaF9jb21wbGlhbmNlX3Jlc3RyaWN0ZWRcIixcInJlYWQ6Y21lZGlhX2Vjb3N5c3RlbV90ZXN0aW5nXCIsXCJyZWFkOjNkX2FkbWluXCIsXCJ3cml0ZTozZF9hZG1pblwiXSxcInN1cHBsaWVyc1wiOntcIjU4MDc5XCI6W1wicmVhZDpleHRyYW5ldF91c2VyXCIsXCJyZWFkOmludmVudG9yeVwiLFwid3JpdGU6aW52ZW50b3J5XCIsXCJyZWFkOmNhdGFsb2dcIixcIndyaXRlOmNhdGFsb2dcIixcInJlYWQ6cHVyY2hhc2Vfb3JkZXJcIixcIndyaXRlOnB1cmNoYXNlX29yZGVyXCIsXCJyZWFkOmxhYmVsX2dlbmVyYXRpb25fZXZlbnRcIixcInZlcmlmaWVkOnNoaXBwaW5nXCIsXCJ2ZXJpZmllZDphZHZhbmNlX3NoaXBfbm90aWNlXCIsXCJyZWFkOnByb2R1Y3RfY2xhc3NcIixcInJlYWQ6bWFudWZhY3R1cmVyXCIsXCJyZWFkOmludm9pY2VcIixcIndyaXRlOmludm9pY2VcIixcInJlYWQ6c3VwcGxpZXJfcmV2ZW51ZVwiLFwicmVhZDpzaG9wX3RoZV9sb29rXCIsXCJ3cml0ZTpkcm9wX3NoaXBfaW52b2ljZVwiLFwidmVyaWZpZWQ6ZHJvcF9zaGlwX2ludm9pY2VcIixcInJlYWQ6Y2FzdGxlZ2F0ZV9zY19vcmRlcnNcIixcIndyaXRlOmNhc3RsZWdhdGVfc2Nfb3JkZXJzXCIsXCJyZWFkOnNlcnZpY2VfYm9va2luZ1wiLFwid3JpdGU6c2VydmljZV9ib29raW5nXCIsXCJ2ZXJpZmllZDpzZXJ2aWNlX2Jvb2tpbmdcIixcInJlYWQ6ZmluYW5jZV9wYXJ0bmVyX2hvbWVcIixcIndyaXRlOmZpbmFuY2VfcGFydG5lcl9ob21lXCJdfX0iLCJhdWQiOiJodHRwczpcL1wvYXBpLndheWZhaXIuY29tXC8iLCJzdWIiOiJncDA4S0MwajNNMEJIaEY3Y0R4Q2xYak5TTU45dWtzV0BjbGllbnRzIiwiaXNzIjoiaHR0cHM6XC9cL3BhcnRuZXJzLndheWZhaXIuY29tXC8iLCJleHAiOjE3MTI0NzU4NDh9.JqtdIs_DZH9Nu4l5El43piMQG9haCaU_w1aP4h7qEZI; _dd_s=rum=1&id=f342c189-e367-4353-8d52-a4fffd35fe92&created=1712468642331&expire=1712469551632; _ga_W5CBQ28KZV=GS1.2.1712463715.64.1.1712468651.35.0.0'}

url = "https://partners.wayfair.com/a/finance/payment/payment_details/get_sap_paid_invoices"
headers =  {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36","Connection": "close"}
row_data_file = "row_data_file.csv"
fields = ['invoiceNumber', 'invoiceCurrency', 'invoiceDate', 'invoiceAmount', 'paymentId', 'paymentDate', 'invoiceType', 'invoiceDetail', 'invoiceSource',"isNo"]

def get_po_num(local_file,sheet_name):
    workbook = load_workbook(local_file)
    sheet = workbook[sheet_name]
    #values_only=True 是指在访问或处理数据时只返回数值，而不返回索引或其他元数据
    for row in sheet.iter_rows(values_only=True):
        yield row[0]
        

def readTxt(process_file):
    if not os.path.exists(process_file):
        with open(process_file, "w") as f:
            f.write('')
            
    with open(process_file, "r") as f:
        data = f.read().splitlines()
        return set(data) 
     
def get_number(tem_number):
    if not tem_number:
        return False  
    
    number = ''
    for i in tem_number:
        if not i.isdigit():
            continue
        number = number + i
        if len(number) == 9:
            return number
        
    return False

def random_sleep(min_time, max_time):
    sleep_time = random.uniform(min_time, max_time)
    time.sleep(sleep_time)
def get_voucher_number():
 # 加载 Excel 文件
    workbook = load_workbook(voucher_number_file)
    # 选择第一个工作表
    sheet =  workbook[sheet_name_vo_num]
    return  set([str(int(cell.value)) if type(cell.value is str) else cell.value  for cell in sheet['A'][1:] if cell.value])
    
    
list_number = get_po_num(local_file,sheet_name)
process_num = readTxt("进度.txt")

voucher_number_lst = get_voucher_number()
  
for po_number in list_number:
    
    if po_number in process_num:
        print(f"查询：{po_number}，跳过")
        continue
        
   
    if not get_number(po_number):
        continue
        
    data = ["falg"]
    count =0 
    payload['poNumber'] = po_number
    while len(data) != 0:
        count = count + 1
        if count > 5:
               
            with open("server_file.csv","a+",newline = '') as csv_f:
                csv.writer(csv_f).writerow([po_number,"0"]) 
                
            with open("进度.txt", "a+") as f:
                f.write(f"{po_number}\n") # 将内容追加到到文件尾部
              
            break
        
        try:
            res = requests.post(url,data=json.dumps(payload),headers=headers,cookies =cookies )
            random_sleep(1,2)
             
            # 使用BeautifulSoup解析HTML内容
            soup = BeautifulSoup(res.text, 'html.parser')

            div_content = soup.find('div', class_='maincontent wfe_container').text.strip()
            data = json.loads(div_content)["data"]["paidInvoices"]
            
            if  len(data) != 0:
                break
        
        except Exception as e :
            print("异常网络，重试",e.__traceback__.tb_lineno,e)
            time.sleep(60)
  
            
    with open(row_data_file,"a+",newline = '') as csv_f:
            csv.writer(csv_f).writerow([po_number]) 
            
    invoice_amounts = [] 
    withhold = 0              
    with open(row_data_file, 'a+', newline='') as csvfile:
        for invoice in data:   
            writer = csv.DictWriter(csvfile, fieldnames=fields)
            invoice['paymentId'] = str(invoice['paymentId'])
            invoice_amount = invoice['invoiceAmount']
            if invoice['paymentId'] not in voucher_number_lst:
                invoice['isNo']  = "No"
                
                
            if invoice_amount < 0:
                withhold = 1
    
            invoice_amounts.append(invoice_amount)
            writer.writerow(invoice)
            
    sum_money = sum([Decimal(str(data)) for data in   invoice_amounts if data] )  
    
   
    
    with open("server_file.csv","a+",newline = '') as csv_f:
        if withhold:
            csv.writer(csv_f).writerow([po_number,str(sum_money),"扣款"]) 
        else:
            csv.writer(csv_f).writerow([po_number,str(sum_money)]) 
    
          
    with open("进度.txt", "a+") as f:
        f.write(f"{po_number}\n") # 将内容追加到到文件尾部
        
    print(f"数据已成功写入CSV文件：{po_number}")
    
    

