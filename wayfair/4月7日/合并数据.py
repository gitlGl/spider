import csv,os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook,Workbook

from openpyxl.utils import column_index_from_string
current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  

server_file = "server_file.csv"


def get_po_num(local_file,sheet_name,coulum):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    sheet = sheet = workbook[sheet_name]
    if coulum == "C":
        lst = []
        for cell in sheet[coulum]:
            if type(cell.value) is str:
                lst.append(None)
                continue
            if cell.value is None:
                lst.append(None)
                continue
            lst.append(cell.value)
        return lst
        
    if coulum == "A":
        return [ cell.value if cell.value else cell.value  for cell in sheet[coulum] ]
        

# 获取CSV文件数据
def get_data(server_file):
    with open(server_file, 'r') as file:
        reader = csv.reader(file)
        
        list_data = []
        for row in reader:
            if len(row)<2:
                print(row)
            list_data.append(row)
        return list_data
    
def mer(colum_num):
    server_datas = get_data(server_file)
    local_datas = get_po_num(local_file,sheet_name,"C")
    list_numer = get_po_num(local_file,sheet_name,"A")
    workbook = load_workbook(local_file)
    sheet = workbook[sheet_name]
    
    green_list = []
    for server_data in  server_datas:
        # 指定行列插入数据
        index = list_numer.index(server_data[0]) + 1
        server_data__ =  float(server_data[1])
        sheet.cell(row=index, column=colum_num, value=server_data__)
        
        fill_green = PatternFill(fill_type='solid', fgColor='2edfa3')  # 绿色填充
        
        
        if len(server_data)>2:
            green_list.append([list_numer[index - 1],local_datas[index -1],server_data__])
            
            for cell in sheet[index]:
                cell.fill =  fill_green
              
            continue

        
        if local_datas[index -1] is None:
            continue
         
        if server_data__ ==  0 :
            green_list.append([list_numer[index - 1],local_datas[index -1],server_data__])
            
            for cell in sheet[index]:
                cell.fill = fill_green
                
            continue
    
        if server_data__ < local_datas[index -1]*0.905 or server_data__  > local_datas[index -1]*0.915:
            green_list.append([list_numer[index - 1],local_datas[index -1],server_data__])
            
            for cell in sheet[index]:
                cell.fill = fill_green
                      
    workbook.save(local_file)
    workbook.close()
    
    # 创建一个新的工作簿
    workbook = Workbook()

    # 获取当前活动的工作表
    sheet = workbook.active
    # 将数据写入工作表
    for row in green_list:
        sheet.append(row)

    # 保存工作簿
    workbook.save(filename='绿色.xlsx')
    workbook.close()

    

    
def clear_format(local_file,colum):
   
    workbook = load_workbook(local_file)

    # 选择要处理的表格
    worksheet = workbook[sheet_name]  # 替换为实际的工作表名称

    # 遍历D列的每个单元格
    for cell in worksheet[colum]:
        # 清除单元格的格式
        cell.number_format = 'General'

    # 保存修改后的文件
    workbook.save(local_file)
        
if "__main__" == __name__:
    colum = "E"
    local_file = "12-1000 测试.xlsx"
    sheet_name = "12-1000"

    clear_format(local_file,colum)
    colum_num = column_index_from_string(colum) 
    mer( colum_num)
    print("合并完成")
    



