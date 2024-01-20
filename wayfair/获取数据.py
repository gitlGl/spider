
from playwright.sync_api import Playwright, sync_playwright
from openpyxl import load_workbook
import time,os,random,csv,math
from decimal import Decimal

current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  

def random_sleep(min_time, max_time):
    sleep_time = random.randint(min_time, max_time)
    time.sleep(sleep_time)
    


def readTxt(process_file):# 读取已下载的公司代码
    if not os.path.exists(process_file):
        with open(process_file, "w") as f:
            f.write('')
    with open(process_file, "r") as f:
        data = f.read().splitlines()
        return set(data) 
     
def get_po_num(local_file,sheet_name):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    # 选择第一个工作表
    sheet = sheet = workbook[sheet_name]
    return [str(cell.value) for cell in sheet['A'] if cell.value ]

def get_number(tem_number):
    number = ''
    for i in tem_number:
        if i == "_" or i == "-":
            break
        if not i.isdigit():
            continue
        number = number + i
    return number
  
  
#   with open(detail_data_file,"a+",newline = '') as csv_f:
#                                 csv.writer(csv_f).writerow([po_number,0])
def clear_row_data(local_file,sheet_name,row_data_file,
                   detail_data_file,server_file ):
    row = get_po_num(local_file,sheet_name)
    list_number = [i[2:] for i in row]
    data_dic = {}
    for i in list_number:
        data_dic[i] = []
      
    row_data_list = get_row_data(row_data_file)

    for i  in row_data_list[:]:
        lst = data_dic.get(i[5],None)
        if lst is None:
            row_data_list.remove(i)
            continue
        data_dic[i[5]].append(i)
        row_data_list.remove(i)
        
    for key,value in data_dic.items():
        sum_money = sum([Decimal(data[-2][1:].replace(",","")) for data in  value if data] )
        sum_money = float(sum_money)
        
        with open(server_file,"a+",newline = '') as csv_f:
            csv.writer(csv_f).writerow([key,str(sum_money)]) 
                                
        with open(detail_data_file,"a+",newline = '') as csv_f:
            csv.writer(csv_f).writerow([key,str(sum_money)])
            for i in value:
                csv.writer(csv_f).writerow(i)

    
        
    
# 获取CSV文件数据
def get_row_data(row_data_file):
    with open(row_data_file, 'r') as file:
        reader = csv.reader(file)
        
        list_data = []
        for row in reader:
            if len(row)<2:
                print(row)
            list_data.append(row)
        return list_data   
        
    
    
def get_set_data(fiel_path):
        if not os.path.exists(fiel_path):
            with open(fiel_path, "w") as f:
                f.write('')
        set_ = set()
        with open(fiel_path, 'r') as file:
            reader = csv.reader(file)
        
            for row in reader:
                tem_number = get_number(''.join(row))
                set_.add(tem_number)
            return set_
          
def run(playwright: Playwright,sheet_names) -> None:
    browser = playwright.firefox.launch(headless=True)
    # 加载本地cookies，免登陆
    context = browser.new_context(storage_state="state.json")
    # 打开页面继续操作
    page = context.new_page()
    for sheet_name in sheet_names:
        server_file = f"{sheet_name}.csv"
       
        process_file = f"{sheet_name}.txt"
        row_data_file = f"{sheet_name}-row_data.csv"
        detail_data_file = f"{sheet_name}-detail.csv"
        process_num = readTxt(process_file)
        list_number = get_po_num(local_file,sheet_name)
        set_data = get_set_data(row_data_file)
        
       
        
        for po_number in list_number[:]:
            if po_number in process_num:
                list_number.remove(po_number)
                print(f"查询：{po_number}，跳过")
    
       
    
        while True:
            try:
                page.goto('https://partners.wayfair.com/v/finance/payment/payments_summary/display')
                page.wait_for_load_state("networkidle")
                break
            except Exception as e :
                print("异常网络，重试",e.__traceback__.tb_lineno,e)
                random_sleep(1,2)
    
    
        for po_number in list_number:
            count = 0
            while True:
                try:
                    page.locator('input[name="poNumber"]').clear()
                    page.locator('input[name="poNumber"]').fill(po_number)
                    page.locator('button:text("Search")').click()
                    page.wait_for_load_state("networkidle")
                    
                    div_rt_tbody = page.query_selector('div.rt-tbody')  
                    div_rt_tr_groups = div_rt_tbody.query_selector_all('div.rt-tr-group')
                    
                    tt = []  
                    flag = 0
                    for div_rt_td in div_rt_tr_groups:
                        div_datas = div_rt_td.query_selector_all('div.rt-td')
                        data_ = []
                        for item in div_datas:
                        
                            text = item.text_content()
                            if text != "\xa0":
                                data_.append(text)
                                tt.append(text)
                        if data_:
                            if len(data_)>3:
                                tem_number = get_number(data_[2])
                                if po_number[2:] == tem_number:
                                    flag = 1
                                    
                                data_.append(tem_number)
                                string = "".join(data_)
                                if string not in set_data:
                                    set_data.add(string)
                                    with open(row_data_file,"a+",newline = '') as csv_f:
                                        
                                        csv.writer(csv_f).writerow(data_)
                
                    if not tt or not flag:
                        count = count +1
                        print(f" {po_number} count:",count)
                        if count > 100:
                            count = 0
                            break
                        random_sleep(1,2)
                        continue
                
                    with open(process_file, "a") as f:
                        f.write(f"{po_number}\n") # 将内容追加到到文件尾部
                    print(f"进度{po_number}")
                    random_sleep(1,2)
                    break                  
                
                except Exception as e:
                    print("异常网络，重试",e.__traceback__.tb_lineno,e,e.__traceback__.tb_frame)
                    random_sleep(1,2)
                    page.locator('input[name="poNumber"]').clear()
                    
        with open("sheet_name.txt", "a+",encoding="utf8") as f:
            f.seek(0)
            data = f.read().splitlines()
             
            if not sheet_name in data:
                print(f"{sheet_name}数据收集完成,正在清洗数据")
                clear_row_data(local_file,sheet_name,row_data_file,
                        detail_data_file,server_file )
                print(f"{sheet_name}清洗完成")
                
                f.seek(0,2)
                f.write(f"{sheet_name}\n") # 将内容追加到到文件尾部
                print(f"进度{sheet_name}")
                
            
    
       
    context.close()
    browser.close()


if "__main__" == __name__:
    local_file = "测试.xlsx"
    workbook = load_workbook(local_file)
    sheet_names = workbook.sheetnames
    
    with sync_playwright() as playwright:
        run(playwright,sheet_names)