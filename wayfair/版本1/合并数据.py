import csv,os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  


def get_po_num(local_file,sheet_name,coulum):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    sheet = sheet = workbook[sheet_name]
    if coulum == "C":
        return [ float(cell.value) if  cell.value else 0 for cell in sheet[coulum] ]
    if coulum == "A":
        return [ cell.value  for cell in sheet[coulum] ]
        

# 获取CSV文件数据
def get_data(server_file):
    with open(server_file, 'r') as file:
        reader = csv.reader(file)
        
        list_data = []
        for row in reader:
            if len(row)<2:
                print(row)
            list_data.append(float(row[1]))
        return list_data
def mer(sheet_names):
   for sheet_name in sheet_names:
        server_file = f"{sheet_name}.csv"

        server_datas = get_data(server_file)
        local_datas = get_po_num(local_file,sheet_name,"C")
        list_number =  get_po_num(local_file,sheet_name,"A")
        workbook = load_workbook(local_file)
        sheet = workbook[sheet_name]

        for index,(server_data,local_data) in enumerate(zip(server_datas,local_datas),start=1):
            # 指定行列插入数据
            if list_number[index-1] is None:
                row =  index + 1
            else :
                row = index
            sheet.cell(row=row, column=4, value=server_data)
            
            if server_data < local_data*0.91:
                # 创建填充对象并设置颜色
                fill = PatternFill(fill_type='solid', fgColor='2edfa3')  # 绿色填充
                #填充行为绿色
                for cell in sheet[index]:
                    cell.fill = fill
                # sheet.cell(row=index, column=3).fill = fill
        
        workbook.save(local_file)
        workbook.close()
def readTxt(sheet_name_file):# 读取已下载的公司代码
    if not os.path.exists(sheet_name_file):
        with open(sheet_name_file, "w",encoding="utf8") as f:
            f.write('')
    with open(sheet_name_file, "r",encoding="utf8") as f:
        data = f.read().splitlines()
        return set(data)        
if "__main__" == __name__:
    local_file = "测试.xlsx"
    sheet_names = readTxt("sheet_name.txt")
    mer(sheet_names)
    print("合并完成")
    



