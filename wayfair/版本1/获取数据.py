
from playwright.sync_api import Playwright, sync_playwright
from openpyxl import load_workbook
import time,os,random,csv
from decimal import Decimal

current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  

def random_sleep(min_time, max_time):
    sleep_time = random.uniform(min_time, max_time)
    time.sleep(sleep_time)
    


def readTxt(process_file):# 读取已下载的公司代码
    if not os.path.exists(process_file):
        with open(process_file, "w") as f:
            f.write('')
    with open(process_file, "r") as f:
        data = f.read().splitlines()
        return set(data) 
     
def get_po_num(local_file,sheet_name):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    # 选择第一个工作表
    sheet = sheet = workbook[sheet_name]
    return [str(cell.value) for cell in sheet['A'] if cell.value ]

def get_number(tem_number):  
    
    number = ''
    for i in tem_number:
        if not i.isdigit():
            continue
        number = number + i
        if len(number) == 9:
            return number
    return number
   
    
def run(playwright: Playwright,sheet_names) -> None:
    
    for sheet_name in sheet_names:
        server_file = f"{sheet_name}.csv"
        process_file = f"{sheet_name}.txt"
        detail_data_file = f"{sheet_name}-detail.csv"
        process_num = readTxt(process_file)
        list_number = get_po_num(local_file,sheet_name)
        
        for po_number in list_number[:]:
            if po_number in process_num:
                list_number.remove(po_number)
                print(f"查询：{po_number}，跳过")
    
        browser = playwright.firefox.launch(headless=True)
        # 加载本地cookies，免登陆
        context = browser.new_context(storage_state="state.json")
        # 打开页面继续操作
        page = context.new_page()
    
        while True:
            try:
                page.goto('https://partners.wayfair.com/v/finance/payment/payments_summary/display')
                page.wait_for_load_state("networkidle")
                break
            except Exception as e :
                print("异常网络，重试",e.__traceback__.tb_lineno,e)
                random_sleep(0.6,1)
    
    
        for po_number in list_number:
            count = 0
            flag =[]
            
            while True:
                try:
                    page.locator('input[name="poNumber"]').clear()
                    page.locator('input[name="poNumber"]').fill(po_number)
                    page.locator('button:text("Search")').click()
                    page.wait_for_load_state("networkidle")
                    random_sleep(0.8,1)
                    
                    div_rt_tbody = page.query_selector('div.rt-tbody')  
                    div_rt_tr_groups = div_rt_tbody.query_selector_all('div.rt-tr-group')
                    
                    tt = []  
                    data_list = []
                    for div_rt_td in div_rt_tr_groups:
                        div_datas = div_rt_td.query_selector_all('div.rt-td')
                        data_ = []
                        for index ,item in enumerate( div_datas):
                            if index == 2:
                                tem_text = item.text_content()
                                print(tem_text)
                                tem_number = get_number(tem_text)
                                print("tem_number:",tem_number," ",po_number[2:])
                                if tem_number != po_number[2:]:
                                    flag.append(False)
                                else: flag.append(True)
                                    
                                
                            text = item.text_content()
                            if text != "\xa0":
                                data_.append(text)
                                tt.append(text)
                        data_list.append(data_)
                    
                
                    if all(flag):
                        flag = []
                    else:
                        flag = []
                        continue
                                
                    if tt:
                        
                        sum_money = sum(Decimal(data[-1][1:].replace(",","")) for data in  data_list if data )
                        sum_money = float(sum_money)
                        print("总金额",sum_money)
                        print(data_list)
                                        
                        with open(server_file,"a+",newline = '') as csv_f:
                                csv.writer(csv_f).writerow([po_number,str(sum_money)]) 
                                
                        with open(detail_data_file,"a+",newline = '') as csv_f:
                            csv.writer(csv_f).writerow([po_number,str(sum_money)])
                            for data in data_list:
                                csv.writer(csv_f).writerow(data)
                                    
                        
                        with open(process_file, "a") as f:
                            f.write(f"{po_number}\n") # 将内容追加到到文件尾部
                        print(f"进度{po_number}")
                        
                       
                        break
                        
                    else:
                        if count > 100 :
                            with open(server_file,"a+",newline = '') as csv_f:
                                csv.writer(csv_f).writerow([po_number,0])
                                    
                            with open(detail_data_file,"a+",newline = '') as csv_f:
                                csv.writer(csv_f).writerow([po_number,0])

                            print(f"进度{po_number}")
                            count = 0
                            break
                            
                        count = count + 1
                      
                        continue 
                    
                except Exception as e:
                    print("异常网络，重试",e.__traceback__.tb_lineno,e)
                   
                    page.locator('input[name="poNumber"]').clear()
        
        with open("sheet_name.txt", "r+",encoding="utf8") as f:
             data = f.read().splitlines()
                
        with open("sheet_name.txt", "a+",encoding="utf8") as f:
            if not sheet_name in data:
                f.write(f"{sheet_name}\n") # 将内容追加到到文件尾部
                print(f"进度{sheet_name}")
                        
            
        
        print(f"{sheet_name}数据收集完成")
        context.close()
        browser.close()


if "__main__" == __name__:
    local_file = "测试.xlsx"
    workbook = load_workbook(local_file)
    sheet_names = workbook.sheetnames
    
    with sync_playwright() as playwright:
        run(playwright,sheet_names)