from playwright.sync_api import Playwright, sync_playwright
from openpyxl import load_workbook
import time,os,csv

def readTxt(process_file):# 读取已下载的公司代码
    if not os.path.exists(process_file):
        with open(process_file, "w") as f:
            f.write('')
            
    with open(process_file, "r") as f:
        data = f.read().splitlines()
        return set(data) 
    
def get_po_num(local_file,sheet_name):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    # 选择第一个工作表
    sheet = sheet = workbook[sheet_name]
    data = []

    for cell in sheet['A']:
        value = cell.value
        if type(value)is str and len(value) == 11:
            data.append(value)
    return data


def check_num(page,num):

    div_rt_tbody = page.query_selector('div.rt-tbody')  
    div_rt_tr_groups = div_rt_tbody.query_selector_all('div.rt-tr-group')

    if len(div_rt_tr_groups)!= num:
        return False
   
    for  div_rt_td in div_rt_tr_groups:
        div_datas = div_rt_td.query_selector_all('div.rt-td')
        for item in div_datas:
            text = item.text_content()
            if text == "\xa0" or  text=="":
                return False
    return True     

     
def get_text(page):
 
    div_rt_tbody = page.query_selector('div.rt-tbody')  
    div_rt_tr_groups = div_rt_tbody.query_selector_all('div.rt-tr-group')
    data_list = []
    for  div_rt_td in div_rt_tr_groups:
        data = []
        div_datas = div_rt_td.query_selector_all('div.rt-td')
        for index,item in enumerate( div_datas):
            text = item.text_content()
            if text != "\xa0" and  text!="":
                if index == 0:
                    data.append(text[:11])

                if index == 3 and data:
                    data.append(text)
            else:
                return False

        if len(data) == 2:
            data_list.append(data)
   
    else:
        return data_list

def to_page(page):
    while True:
        try:
            page.goto('https://partners.wayfair.com/d/outbound-orders/po')
            page.wait_for_load_state("networkidle")
            break
        except Exception as e :
            print("异常网络，重试",e.__traceback__.tb_lineno,e)
            time.sleep(2)           

def run(playwright: Playwright,sheets) -> None:
    browser = playwright.firefox.launch(headless=False)
    # 加载本地cookies，免登陆
    context = browser.new_context(storage_state="state.json")
    # 打开页面继续操作
    page = context.new_page()
     
    for sheet_name in sheet_names:
      
        list_number = get_po_num(local_file,sheet_name)
        process_file = f"{sheet_name}.txt"
        process_num = readTxt(process_file)

        row_data_file = f"{sheet_name}-row_data.csv"
               
        for po_number in list_number[:]:
            if po_number in process_num:
                list_number.remove(po_number)
                print(f"查询：{po_number}，跳过")
    

        to_page(page)
        process = []
        count = 1
        retry = 0
        for index,po_number in enumerate(list_number):
            while True:
                try:
                    print(f"正在查询{po_number}")
                    page.locator('[data-tag-default="components_search_bar_TextInput"]').clear()
                    page.locator('[data-tag-default="components_search_bar_TextInput"]').fill(po_number)
                    page.locator('[aria-label="Search"]').click()
                    page.wait_for_load_state("networkidle")
                    time.sleep(3)
                    if check_num(page,count):
                        process.append(po_number)
                        count = count + 1
                        break

                    else:
                        print(f"重试次数{po_number}:{retry}")
                        time.sleep(2)
                        retry = retry + 1
                        if retry == 10 :
                            break
                        continue
                    
                except Exception as e:
                    print("异常网络，重试",e.__traceback__.tb_lineno,e,e.__traceback__.tb_frame)
            if count == 11:
                time.sleep(5)
                data_list = get_text(page)
                with open(row_data_file,"a+",newline = '') as csv_f1:
                        for xx in data_list:
                            csv.writer(csv_f1).writerow(xx)

                with open(process_file, "a") as f:
                    for x in process:
                        f.write(f"{x}\n") # 将内容追加到到文件尾部
                        print(f"进度{x}")

                to_page(page)
                count = 1
                process = []

            if index + 1 == len(list_number):
                time.sleep(5)
                data_list = get_text(page)
                with open(row_data_file,"a+",newline = '') as csv_f1:
                        for xx in data_list:
                            csv.writer(csv_f1).writerow(xx)

                with open(process_file, "a") as f:
                    for x in process:
                        f.write(f"{x}\n") # 将内容追加到到文件尾部
                        print(f"进度{x}")

                to_page(page)
                count = 1

        with open("sheet_name.txt", "a+",encoding="utf8") as f:
            f.seek(0)
            data = f.read().splitlines()
             
            if not sheet_name in data:
                f.seek(0,2)
                f.write(f"{sheet_name}\n") # 将内容追加到到文件尾部
                print(f"进度{sheet_name}")

                     
    context.close()
    browser.close()
            

if "__main__" == __name__:
    local_file = "2.9 CK payment.xlsx"
    workbook = load_workbook(local_file)
    sheet_names = workbook.sheetnames
    with sync_playwright() as playwright:
        run(playwright,sheet_names)