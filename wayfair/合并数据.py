import csv,os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

from openpyxl.utils import column_index_from_string


current_file_path = os.path.abspath(__file__)
os.chdir(os.path.dirname(current_file_path))  


def get_po_num(local_file,sheet_name,coulum):
    # 加载 Excel 文件
    workbook = load_workbook(local_file)
    sheet = sheet = workbook[sheet_name]
    if coulum == "C":
        lst = []
        for cell in sheet[coulum]:
            if type(cell.value) is str:
                lst.append(None)
                continue
            if cell.value is None:
                lst.append(None)
                continue
            lst.append(cell.value)
        return lst
        
    if coulum == "A":
        return [ cell.value[2:] if cell.value else cell.value  for cell in sheet[coulum] ]
        

# 获取CSV文件数据
def get_data(server_file):
    with open(server_file, 'r') as file:
        reader = csv.reader(file)
        
        list_data = []
        for row in reader:
            if len(row)<2:
                print(row)
            list_data.append(row)
        return list_data
def mer(sheet_names,colum_num):
   for sheet_name in sheet_names:
        server_file = f"{sheet_name}.csv"

        server_datas = get_data(server_file)
        local_datas = get_po_num(local_file,sheet_name,"C")
        list_numer = get_po_num(local_file,sheet_name,"A")
        workbook = load_workbook(local_file)
        sheet = workbook[sheet_name]

        for server_data in  server_datas:
            # 指定行列插入数据
            index = list_numer.index(server_data[0]) + 1
            server_data__ =  float(server_data[1])
            sheet.cell(row=index, column=colum_num, value=server_data__)
            
            fill_green = PatternFill(fill_type='solid', fgColor='2edfa3')  # 绿色填充
            fill_red = PatternFill(fill_type='solid', fgColor='ed5a65')  # 绿色填充
            
            if len(server_data)>2:
                 #填充行为红色
                 
                for cell in sheet[index]:
                    cell.fill = fill_red
                continue
          
            if local_datas[index -1] is None:
                continue
           
            if server_data__ < local_datas[index -1]*0.905:
                # 创建填充对象并设置颜色
                #填充行为绿色
                for cell in sheet[index]:
                    cell.fill = fill_green
                # sheet.cell(row=index, column=3).fill = fill
        
        workbook.save(local_file)
        workbook.close()
def readTxt(sheet_name_file):# 读取已下载的公司代码
    if not os.path.exists(sheet_name_file):
        with open(sheet_name_file, "w",encoding="utf8") as f:
            f.write('')
    with open(sheet_name_file, "r",encoding="utf8") as f:
        data = f.read().splitlines()
        return set(data)   
    
def clear_format(local_file,sheet_names,colum):
    for sheet_name in sheet_names:
        workbook = load_workbook(local_file)

        # 选择要处理的表格
        worksheet = workbook[sheet_name]  # 替换为实际的工作表名称

        # 遍历D列的每个单元格
        for cell in worksheet[colum]:
            # 清除单元格的格式
            cell.number_format = 'General'

        # 保存修改后的文件
    workbook.save(local_file)
        
if "__main__" == __name__:
    colum = "E"
    local_file = "Testing11.20-11.30.xlsx"
    sheet_names = readTxt("sheet_name.txt")
    clear_format(local_file,sheet_names,colum)
    colum_num = column_index_from_string(colum) 
    mer(sheet_names, colum_num)
    print("合并完成")
    



